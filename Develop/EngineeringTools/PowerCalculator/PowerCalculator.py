##################################################
# Program: PowerCalculator
##################################################

##################################################
# System Module Imports
##################################################
from enum import Enum
import queue
from xml.etree.ElementTree import Element, SubElement, Comment, ElementTree, tostring
##################################################
# Add-In Module Imports
##################################################
from Drivers.PyQt_activity import BaseActivity, BaseActivityWindow
from Drivers.string_funcs import get_string_from_stream
from Drivers.xml_prettyprint import prettify
import openpyxl
from openpyxl.utils import column_index_from_string
from PyQt5 import QtWidgets, QtCore
from Qt.UiFunctions.open import get_file_with_dialogue
from Qt.UiFunctions.save import save_file_with_dialogue
from Qt.UiFunctions.status_bar import set_status_bar
##################################################
# Local Module Imports
##################################################

##################################################
# Constant Variable Definitions
##################################################

##################################################
# Global Variable Definitions
##################################################


class Messages(Enum):  # enum classes exist as a "template" of sorts to allow easy-to-implement / read code.
    LOAD_CALC = 10
    SAVE_RESULTS = 20


class PowerCalculatorWindow(BaseActivityWindow):
    _set_browser_text = QtCore.pyqtSignal(str)
    _set_save_enable_state = QtCore.pyqtSignal(bool)

    def __init__(self,
                 queue_ref: [queue.Queue],
                 parent=None):
        super().__init__(queue_ref=queue_ref,
                         parent=parent)

    def __setup_layout__(self):
        # setup layout
        self._v_layout = QtWidgets.QVBoxLayout(self)
        self._v_layout.setContentsMargins(5, 5, 5, 5)
        self._v_layout.setSpacing(5)
        self._h_layout = QtWidgets.QHBoxLayout(self)
        self._h_layout.setContentsMargins(5, 5, 5, 5)

        # create objects
        self._title_bar = QtWidgets.QLabel("Indicon Power Calculator XML Preview")
        self._title_bar.setAlignment(QtCore.Qt.AlignCenter)
        self._title_bar.setStyleSheet("background-color: white;color: black")

        self._text_browser = QtWidgets.QTextBrowser(self)
        self._text_browser.setDocumentTitle("XML Preview Browser")

        self._load_calc_pushbutton = QtWidgets.QPushButton("Load Calc")
        self._load_calc_pushbutton.clicked.connect(lambda: self._queue_ref.put(Messages.LOAD_CALC))

        self._save_to_xml_push_button = QtWidgets.QPushButton("Save To XML")
        self._save_to_xml_push_button.clicked.connect(lambda: self._queue_ref.put(Messages.SAVE_RESULTS))
        self._save_to_xml_push_button.setEnabled(False)

        # assign objects to layout
        self._v_layout.addWidget(self._title_bar)
        self._v_layout.addWidget(self._text_browser)
        self._v_layout.addLayout(self._h_layout)

        self._h_layout.addWidget(self._load_calc_pushbutton)
        self._h_layout.addWidget(self._save_to_xml_push_button)

        self.setLayout(self._v_layout)

    def __bind_connections__(self):
        self._set_browser_text.connect(self.__update_browser_text__)
        self._set_save_enable_state.connect(self.__set_save_enable_button_state__)

    def update_browser_text(self, text):
        self._set_browser_text.emit(text)

    def __update_browser_text__(self, text):
        self._text_browser.setText(text)

    def set_save_enable_button_state(self, state):
        self._set_save_enable_state.emit(state)

    def __set_save_enable_button_state__(self, state):
        self._save_to_xml_push_button.setEnabled(state)


class PowerCalculator(BaseActivity):
    gui_class = PowerCalculatorWindow

    def __init__(self, gui_ref, queue_ref):
        super().__init__(gui_ref, queue_ref)
        self._system_name = None
        self._panel_name = None
        self._max_amps = None
        self._current_amps = None
        self._vac480_breakers = []
        self._vac120_xfmrs = []
        self._xml_stream = None

    def __run__(self, message):
        match message:
            case Messages.LOAD_CALC:
                self.__load_calc_from_excel__()
            case Messages.SAVE_RESULTS:  # this is just an example, this doesn't really do anything
                some_path = save_file_with_dialogue()
                self.__save_calc_upload_to_xml__(some_path)

    def __load_calc_from_excel__(self):
        # set status bar with new text
        set_status_bar("Loading/Parsing Calculator File")
        # set local variables
        constHeader = 0
        constFooter = 1
        constColumn = 2
        constRow = 3

        # assemble local dicts to make parsing the file easier
        header_keys = {
            'begin': '<',
            'namespace': '@',
            'end': '>'
        }
        header_kw = {
            'header': [f'{header_keys["begin"]}{header_keys["namespace"]}HEADER', f'{header_keys["end"]}', None, None],
            'processor': [f'{header_keys["begin"]}{header_keys["namespace"]}PROCESSOR ', f'{header_keys["end"]}', 1, 1],
            'panel': [f'{header_keys["begin"]}{header_keys["namespace"]}PANEL ', f'{header_keys["end"]}', 1, 1],
            'disconnect': [f'{header_keys["begin"]}{header_keys["namespace"]}DISCONNECT ', f'{header_keys["end"]}', 1, 1],
            'totalAmp': [f'{header_keys["begin"]}{header_keys["namespace"]}TOTALAMP ', f'{header_keys["end"]}', 1, 1],
            '_480breaker': [f'{header_keys["begin"]}{header_keys["namespace"]}480BREAKER ', f'{header_keys["end"]}', 1, 1],
            '_480breakerName': [f'{header_keys["begin"]}{header_keys["namespace"]}480BREAKERNAME ', f'{header_keys["end"]}', 1, 1],
            '_480amps': [f'{header_keys["begin"]}{header_keys["namespace"]}480AMPS ', f'{header_keys["end"]}', 1, 1],
            '_120draw': [f'{header_keys["begin"]}{header_keys["namespace"]}120DRAW ', f'{header_keys["end"]}', 1, 1],
            '_120breaker': [f'{header_keys["begin"]}{header_keys["namespace"]}120BREAKER ', f'{header_keys["end"]}', 1, 1],
            '_120amps': [f'{header_keys["begin"]}{header_keys["namespace"]}120AMPS ', f'{header_keys["end"]}', 1, 1],
            '_480dev': [f'{header_keys["begin"]}{header_keys["namespace"]}480DEV ', f'{header_keys["end"]}', 1, 1],
            '_480devAmps': [f'{header_keys["begin"]}{header_keys["namespace"]}480DEVAMPS ', f'{header_keys["end"]}', 1, 1],
            '_120dev': [f'{header_keys["begin"]}{header_keys["namespace"]}120DEV ', f'{header_keys["end"]}', 1, 1],
            '_120devAmps': [f'{header_keys["begin"]}{header_keys["namespace"]}120DEVAMPS ', f'{header_keys["end"]}', 1, 1],
            '_24dev': [f'{header_keys["begin"]}{header_keys["namespace"]}24DEV ', f'{header_keys["end"]}', 1, 1],
            '_24devAmps': [f'{header_keys["begin"]}{header_keys["namespace"]}24DEVAMPS ', f'{header_keys["end"]}', 1, 1],
            'endHeader': [f'{header_keys["begin"]}{header_keys["namespace"]}ENDHEADER', f'{header_keys["end"]}', None, None],
        }

        # get excel file from qt manager
        excel_file = get_file_with_dialogue(get_env='HOME', file_type_args='.xlsx(*.xlsx)')

        # do not continue unless a file was selected
        if not excel_file:
            return

        # shut off save functionality until we're done
        # this should never really be an issue, as there are no interupts, but i feel this is good practice in case we add interupts later
        self.gui_ref.set_save_enable_button_state(False)

        # read the file into memory
        pdp_calc_wb = openpyxl.load_workbook(excel_file, data_only=True)  # the data only flag will allow us to get values of formulas instead of the formulas themselves
        sheets = pdp_calc_wb.get_sheet_names()

        # look for "title" sheet or " first sheet " in book
        title_sheet: openpyxl.workbook.workbook.Worksheet | None = None
        for sheet in pdp_calc_wb.get_sheet_names():
            if pdp_calc_wb.worksheets.index(pdp_calc_wb[sheet]) == 0:
                title_sheet = pdp_calc_wb[sheet]
                break
        # this should "never" happen as there should always be an index 0, but this is just a catch "in case"
        if not title_sheet:
            return

        # find header to associate value look-ups

        for row in title_sheet.iter_rows():
            if row[0].value == f"{header_kw['header'][constHeader]}{header_kw['header'][constFooter]}":
                PowerCalculator.header_row = row[0].row
                for col in row:
                    for key in header_kw:
                        if key == 'header':  # if we are looking at the header, continue
                            continue
                        if key == 'endHeader':  # if we are looking at the end of the header, break the loop
                            break
                        if col.value.find(header_kw[key][constHeader]) != -1:  # if we found a keyword header then assign the coordinates provided from the template
                            header_kw[key][constColumn], header_kw[key][constRow] = self.__parse_cell_for_header__(col.value)
                            break  # break, we are done if we assigned values
                break  # we found what we need, move on

        # ok, let's parse the file with the header information present
        # until calculator is fixed, we must ignore any 480 devices after the xformer section of the calculator starts
        _ignore_480_devices = False
        for row in title_sheet.iter_rows(min_row=PowerCalculator.header_row + 2, max_row=title_sheet.max_row):  # add 2 to header row to skip A) the header and B) the human readable header (plain text header stuff)

            if row[header_kw['processor'][constColumn]].value is not None or row[header_kw['panel'][constColumn]].value is not None:
                self._system_name = row[header_kw['processor'][constColumn]].value
                self._panel_name = row[header_kw['panel'][constColumn]].value
                self._max_amps = row[header_kw['disconnect'][constColumn]].value
                self._current_amps = row[header_kw['totalAmp'][constColumn]].value
                continue  # do not continue parsing this row

            if row[header_kw['_480breaker'][constColumn]].value:
                _xfmr = row[header_kw['totalAmp'][constColumn]].value.find('Disconnect') if row[header_kw['totalAmp'][constColumn]].value else None  # check to see if this is a transformer or a 480 circuit breaker
                if (not _xfmr) and (_ignore_480_devices is False):  # ignore if we moved into 120 transformer section. Calculator must be fixed before this check can be removed
                    new_breaker = PDP480Breaker()
                    new_breaker.name = row[header_kw['_480breakerName'][constColumn]].value
                    new_breaker.breaker_size = row[header_kw['_480breaker'][constColumn]].value
                    new_breaker.current_amps = row[header_kw['_480amps'][constColumn]].value
                    self._vac480_breakers.append(new_breaker)
                if _xfmr:
                    _ignore_480_devices = True
                    new_xfmr = PDP120Xfmr()
                    new_xfmr.name = row[header_kw['_480breakerName'][constColumn]].value
                    new_xfmr.xfmr_size = row[header_kw['_480breaker'][constColumn]].value
                    new_xfmr.current_amps = row[header_kw['_120draw'][constColumn]].value
                    self._vac120_xfmrs.append(new_xfmr)
                continue  # do not continue parsing this row

            if row[header_kw['_480dev'][constColumn]].value and (_ignore_480_devices is False):  # ignore if we moved into 120 transformer section. Calculator must be fixed before this check can be removed
                new_device = ChildDevice()
                new_device.name = row[header_kw['_480dev'][constColumn]].value
                new_device.amp_draw = row[header_kw['_480devAmps'][constColumn]].value
                self._vac480_breakers[-1].children_devices.append(new_device)
                continue  # do not continue parsing this row

            if row[header_kw['_120breaker'][constColumn]].value:
                new_breaker = PDP120Breaker()
                new_breaker.name = row[header_kw['_120breaker'][constColumn]].value
                new_breaker.current_amps = row[header_kw['_120amps'][constColumn]].value
                self._vac120_xfmrs[-1].children_devices.append(new_breaker)  # append to last eval'd 120 xformer
                continue  # do not continue parsing this row

            if row[header_kw['_120dev'][constColumn]].value:
                new_device = ChildDevice()
                new_device.name = row[header_kw['_120dev'][constColumn]].value
                new_device.amp_draw = row[header_kw['_120devAmps'][constColumn]].value
                self._vac120_xfmrs[-1].children_devices[-1].children_devices.append(new_device)  # append to last eval'd 120 xformer's last eval'd breaker
                continue  # do not continue parsing this row

        # create xml stream of our compiled information
        self.__compile_to_xml__()
        self.gui_ref.update_browser_text(self._xml_stream)
        self.gui_ref.set_save_enable_button_state(True)
        set_status_bar("Loading/Parsing Calculator File - Complete")

    def __save_calc_upload_to_xml__(self, save_path):
        set_status_bar("Saving...")
        if not self._xml_stream:  # we can't save something we don't have KNOW'M'SAYIN?
            return
        save_path_split = save_path.split('.')
        if save_path_split[-1] != 'XML':
            save_path += '.XML'
        # save this XML with the given save path
        with open(save_path, 'w') as f:
            f.write(self._xml_stream)
        set_status_bar("Save Complete")

    def __compile_to_xml__(self):
        set_status_bar("Compiling XML File From Calculator")
        root = Element('Root')  # root of XML tree
        root.set('version', '1.0')
        comment = Comment('Indicon Corporation Compiled Power Distribution Panel XML Generated File')  # root comment
        root.append(comment)

        # panel info
        panel_info = SubElement(root, 'PanelInfo', {'SystemName': self._system_name,
                                                    'PanelName': self._panel_name,
                                                    'MaxPanelAmps': self._max_amps.__str__(),
                                                    'CurrentPanelAmps': self._current_amps.__str__()})

        # 480 breakers
        c480_breakers = SubElement(root, "c480Breakers")
        for breaker in self._vac480_breakers:
            new_breaker = SubElement(c480_breakers, "c480Breaker", {'Name': breaker.name, 'Size': breaker.breaker_size.__str__(), 'Amps': breaker.current_amps.__str__()})
            for child in breaker.children_devices:
                new_child_device = SubElement(new_breaker, "Child_Device", {'Name': child.name, 'Amps': child.amp_draw.__str__()})

        # 120 transformers
        c120_xformers = SubElement(root, "c120Transformers")
        for xformer in self._vac120_xfmrs:
            new_xformer = SubElement(c120_xformers, "c120Transformer", {'Name': xformer.name, 'Amps': xformer.current_amps.__str__()})
            for child in xformer.children_devices:
                new_child_device = SubElement(new_xformer, "c120Breaker", {'Name': child.name, 'Amps': child.current_amps.__str__()})
                for device in child.children_devices:
                    new_device = SubElement(new_child_device, "Child_Device", {'Name': device.name, 'Amps': device.amp_draw.__str__()})

        self._xml_stream = prettify(root)

    @staticmethod
    # CAUTION !!! - This function reduces the 'y' (column) coordinate by 1, as openpyxl seemingly requires the first iter (e.g. row) to be 1 or greater, but itering the second parameter starts from 0...
    # since our power calculator tool runs by row, this is "ok" but not exactly safe
    # if you want the column to be the root "iter" you'll have to add "1" to the initial value or iter directly from openpyxl (e.g. ws.iter_columns())
    def __parse_cell_for_header__(cell_value):
        x = get_string_from_stream(begin="'", end="'", stream=cell_value, trim_start=True, trim_end=True)
        return column_index_from_string(x[0])-1, x[1]


class PDP480Breaker:
    def __init__(self):
        self.name = None
        self.breaker_size = None
        self.current_amps = None
        self.children_devices = []


class PDP120Breaker:
    def __init__(self):
        self.name = None
        self.current_amps = None
        self.children_devices = []


class PDP120Xfmr:
    def __init__(self):
        self.name = None
        self.xfmr_size = None
        self.current_amps = None
        self.children_devices = []


class ChildDevice:
    def __init__(self):
        self.name = None
        self.amp_draw = None

